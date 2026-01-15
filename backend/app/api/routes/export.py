"""
Export API routes for FinCortex.
Provides Excel and PDF export functionality for reimbursements.
"""

import io
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.services.supabase_service import get_supabase_client

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/export/excel")
async def export_reimbursements_excel(
    manager_id: str = Query(..., description="Manager ID to fetch reimbursements"),
    status: Optional[str] = Query(None, description="Filter by status"),
    start_date: Optional[str] = Query(None, description="Filter start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter end date (YYYY-MM-DD)"),
):
    """
    Export reimbursements to Excel format.
    
    Args:
        manager_id: Manager UUID to fetch reimbursements
        status: Optional status filter (pending, approved, rejected)
        start_date: Optional start date filter
        end_date: Optional end date filter
    
    Returns:
        Excel file as streaming response
    """
    try:
        # Import openpyxl here to avoid startup overhead
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        supabase = get_supabase_client()
        
        # Build query
        query = supabase.table("reimbursements").select(
            "reimbursement_id, receipt_code, vendor_name, amount_claimed, status, expense_date, created_at, users(full_name, email)"
        )
        
        # Note: In production, filter by manager's team members
        # For now, we use a simpler approach
        
        if status:
            query = query.eq("status", status)
        if start_date:
            query = query.gte("expense_date", start_date)
        if end_date:
            query = query.lte("expense_date", end_date)
        
        response = query.order("created_at", desc=True).limit(1000).execute()
        reimbursements = response.data or []
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reimbursements"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define headers
        headers = [
            "Receipt Code",
            "Employee Name",
            "Email",
            "Vendor",
            "Amount",
            "Status",
            "Expense Date",
            "Submitted At"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, reimb in enumerate(reimbursements, 2):
            user = reimb.get("users") or {}
            ws.cell(row=row_idx, column=1, value=reimb.get("receipt_code", "N/A"))
            ws.cell(row=row_idx, column=2, value=user.get("full_name", "Unknown"))
            ws.cell(row=row_idx, column=3, value=user.get("email", ""))
            ws.cell(row=row_idx, column=4, value=reimb.get("vendor_name", ""))
            ws.cell(row=row_idx, column=5, value=float(reimb.get("amount_claimed", 0)))
            ws.cell(row=row_idx, column=6, value=reimb.get("status", "pending").title())
            ws.cell(row=row_idx, column=7, value=reimb.get("expense_date", ""))
            ws.cell(row=row_idx, column=8, value=reimb.get("created_at", "")[:10] if reimb.get("created_at") else "")
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Adjust column widths
        column_widths = [15, 20, 30, 25, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"reimbursements_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl: pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/export/pdf")
async def export_reimbursements_pdf(
    manager_id: str = Query(..., description="Manager ID to fetch reimbursements"),
    status: Optional[str] = Query(None, description="Filter by status"),
    start_date: Optional[str] = Query(None, description="Filter start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter end date (YYYY-MM-DD)"),
):
    """
    Export reimbursements to PDF format.
    
    Note: This is a simplified implementation using HTML-to-PDF.
    For production, consider using reportlab or weasyprint.
    """
    try:
        supabase = get_supabase_client()
        
        # Build query
        query = supabase.table("reimbursements").select(
            "reimbursement_id, receipt_code, vendor_name, amount_claimed, status, expense_date, created_at, users(full_name, email)"
        )
        
        if status:
            query = query.eq("status", status)
        if start_date:
            query = query.gte("expense_date", start_date)
        if end_date:
            query = query.lte("expense_date", end_date)
        
        response = query.order("created_at", desc=True).limit(100).execute()
        reimbursements = response.data or []
        
        # Generate HTML report
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Reimbursement Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 20px; }}
                h1 {{ color: #0891b2; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background: #0891b2; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background: #f8fafc; }}
                .total {{ font-weight: bold; margin-top: 20px; }}
                .approved {{ color: #10b981; }}
                .rejected {{ color: #ef4444; }}
                .pending {{ color: #f59e0b; }}
            </style>
        </head>
        <body>
            <h1>FinCortex Reimbursement Report</h1>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total Records: {len(reimbursements)}</p>
            
            <table>
                <thead>
                    <tr>
                        <th>Receipt Code</th>
                        <th>Employee</th>
                        <th>Vendor</th>
                        <th>Amount</th>
                        <th>Status</th>
                        <th>Date</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        total_amount = 0
        for reimb in reimbursements:
            user = reimb.get("users") or {}
            amount = float(reimb.get("amount_claimed", 0))
            total_amount += amount
            status_class = reimb.get("status", "pending").lower()
            
            html_content += f"""
                <tr>
                    <td>{reimb.get('receipt_code', 'N/A')}</td>
                    <td>{user.get('full_name', 'Unknown')}</td>
                    <td>{reimb.get('vendor_name', '')}</td>
                    <td>${amount:.2f}</td>
                    <td class="{status_class}">{reimb.get('status', 'pending').title()}</td>
                    <td>{reimb.get('expense_date', '')}</td>
                </tr>
            """
        
        html_content += f"""
                </tbody>
            </table>
            <p class="total">Total Amount: ${total_amount:.2f}</p>
        </body>
        </html>
        """
        
        # Return HTML (browsers can print to PDF)
        # For actual PDF, install weasyprint: pip install weasyprint
        filename = f"reimbursements_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        return StreamingResponse(
            io.BytesIO(html_content.encode()),
            media_type="text/html",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"PDF export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
